from typing import Union
from fastapi import FastAPI
from openpyxl import Workbook, load_workbook
from pydantic import BaseModel

class Item(BaseModel):
    dummy : str
    dato1 : Union[str, None] = None
    dato2 : Union[str, None] = None
    dato3 : Union[str, None] = None
    dato4 : Union[str, None] = None
    dato5 : Union[str, None] = None


app = FastAPI()


@app.get("/modelo1/{filename}")
def read_item(filename : str):
    filename_complete = f"{filename}.xlsx"
    wb = load_workbook(filename_complete)
    sheet = wb.active
    item = Item(
        dummy = sheet["A1"].value
        ,dato1 = sheet["B1"].value
        # ,dato2 = sheet["C1"].value
        # ,dato3 = sheet["D1"].value
        # ,dato4 = sheet["E1"].value
        # ,dato5 = sheet["F1"].value
    )
    return item


@app.post("/modelo1/{filename}")
def new_file(filename : str, item :Item):
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = item.dummy
    sheet["B1"] = item.dato1
    sheet["C1"] = item.dato2
    sheet["D1"] = item.dato3
    sheet["E1"] = item.dato4
    sheet["F1"] = item.dato5

    filename_complete = f"{filename}.xlsx"
    workbook.save(filename=filename_complete)

    return {"filename": filename_complete}